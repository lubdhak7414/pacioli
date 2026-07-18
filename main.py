"""Pacioli — FastAPI backend with human-in-the-loop approval."""

import asyncio
import contextvars
import logging
import json
import re
import shutil
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path

import aiosqlite
from fastapi import Depends, FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel, field_validator
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from fastapi.responses import JSONResponse

import config
import db
import ai_client
import ledger_engine
import report_engine
from models import Proposal, OperationType

# ── Logging ───────────────────────────────────────────────────────
_request_id: contextvars.ContextVar[str] = contextvars.ContextVar("request_id", default="-")

logging.basicConfig(
    level=getattr(logging, config.LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(request_id)s] %(name)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Inject request_id into log format
class _RequestIdFilter(logging.Filter):
    def filter(self, record):
        record.request_id = _request_id.get("-")
        return True

logging.getLogger().addFilter(_RequestIdFilter())

# ── Rate limiting ──────────────────────────────────────────────────
limiter = Limiter(key_func=get_remote_address)

# ── Input sanitisation ────────────────────────────────────────────
_ROLE_TAGS = re.compile(r"\[(SYSTEM|ASSISTANT|INST|USER)\]", re.IGNORECASE)
_INJECTION = re.compile(
    r"\b(ignore|disregard|forget)\s+(all\s+)?(previous|above|prior)\b",
    re.IGNORECASE,
)


def sanitize_input(text: str) -> str:
    text = _ROLE_TAGS.sub("", text)
    if _INJECTION.search(text):
        logger.warning("Possible prompt-injection attempt detected.")
    return text[: config.MAX_INPUT_LENGTH].strip()


def fiscal_warnings(actions) -> list[str]:
    """Flag insert_row transactions dated outside the configured fiscal year (item 1.3)."""
    seen: set[str] = set()
    warnings: list[str] = []
    for a in actions:
        if a.operation == OperationType.INSERT_ROW and a.values:
            date_val = a.values[0]
            if isinstance(date_val, str) and date_val[:4].isdigit():
                if int(date_val[:4]) != config.FISCAL_YEAR and date_val not in seen:
                    seen.add(date_val)
                    warnings.append(
                        f"Transaction dated {date_val} is outside fiscal year "
                        f"{config.FISCAL_YEAR}."
                    )
    return warnings


def _trim_for_context(history: list[dict], max_chars: int = 600) -> list[dict]:
    """Cap each chat turn before resending it to the model.

    Assistant turns can include whole rendered report tables; re-feeding those
    verbatim bloats the prompt for no benefit. Keep role + a bounded snippet.
    """
    trimmed = []
    for m in history:
        content = m["content"]
        if len(content) > max_chars:
            content = content[:max_chars] + " …[truncated]"
        trimmed.append({"role": m["role"], "content": content})
    return trimmed


def _render_ai_report(report: dict) -> str:
    """Fallback renderer when the report type can't be computed locally."""
    text = f"**{report.get('title', 'Report')}**\n\n"
    for section in report.get("sections", []):
        text += f"### {section.get('heading', '')}\n"
        for line in section.get("lines", []):
            acct = line.get("account", "")
            num = line.get("account_number", "")
            amt = line.get("amount", 0)
            text += f"  - {acct} ({num}): ${amt:,.2f}\n"
        text += "\n"
    text += "\n_Note: figures could not be recomputed from the ledger; shown as estimated by the AI._"
    return text


# ── Lifespan ──────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    Path("data").mkdir(exist_ok=True)
    await db.init_db()
    logger.info("Database initialised.")

    # Seed the live ledger from the committed template on first run. The live
    # file holds real financial data and is gitignored; the template is the
    # clean starting workbook shipped with the repo.
    ledger_p = Path(ledger_engine.LEDGER_PATH)
    template_p = Path(config.LEDGER_TEMPLATE_PATH)
    if not ledger_p.exists() and template_p.exists():
        ledger_p.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_p, ledger_p)
        logger.info("Seeded ledger from template → %s", ledger_p)

    # Backup ledger on startup (keep last 10)
    if ledger_p.exists():
        backups = sorted(Path("data").glob("ledger_backup_*.xlsx"))
        for old in backups[:-10]:
            old.unlink(missing_ok=True)
        backup_name = f"data/ledger_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(ledger_engine.LEDGER_PATH, backup_name)
        logger.info("Ledger backed up → %s", backup_name)

    # Background task: expire stale proposals every 5 min
    cleanup_task = asyncio.create_task(db.cleanup_stale_proposals())

    yield

    cleanup_task.cancel()


# ── App ───────────────────────────────────────────────────────────
app = FastAPI(
    title="Pacioli",
    version="2.0.0",
    lifespan=lifespan,
    description="Human-in-the-loop AI-powered ledger management.",
)
app.state.limiter = limiter

@app.exception_handler(RateLimitExceeded)
async def rate_limit_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(
        status_code=429,
        content={"detail": "Too many requests. Please slow down."},
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

from fastapi.staticfiles import StaticFiles
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.middleware("http")
async def log_requests(request: Request, call_next):
    rid = uuid.uuid4().hex[:8]
    _request_id.set(rid)
    t0 = time.time()
    response = await call_next(request)
    ms = (time.time() - t0) * 1000
    logger.info("<%s %s> %d  %.0fms", request.method, request.url.path,
                response.status_code, ms)
    response.headers["X-Request-ID"] = rid
    return response


# ── Request/Response models ───────────────────────────────────────
class ChatRequest(BaseModel):
    message: str

    @field_validator("message")
    @classmethod
    def not_empty(cls, v):
        if not v.strip():
            raise ValueError("Message cannot be empty")
        return v.strip()


class ChatResponse(BaseModel):
    assistant_message: str
    proposal_id: int | None = None
    proposal_summary: str | None = None


class ProposalDetail(BaseModel):
    id: int
    status: str
    justification: str
    actions: list[dict]
    user_message: str
    created_at: str
    validation_notes: list[str] = []
    highlight_cells: dict[str, list[str]] = {}


class ApprovalResponse(BaseModel):
    success: bool
    message: str
    change_log: list[str] | None = None


# ── Authentication ─────────────────────────────────────────────────
async def require_auth(request: Request):
    """Check X-API-Key header or ?key= query param against APP_PASSWORD.

    When APP_PASSWORD is empty (default), auth is disabled — zero-config for
    local development.
    """
    if not config.APP_PASSWORD:
        return
    key = request.headers.get("X-API-Key") or request.query_params.get("key")
    if key != config.APP_PASSWORD:
        raise HTTPException(
            status_code=401,
            detail="Invalid or missing API key. Set X-API-Key header.",
        )


# ── API Endpoints ─────────────────────────────────────────────────

@app.get("/api/health", summary="System health check")
async def health_check():
    db_ok = False
    try:
        async with aiosqlite.connect(db.DB_PATH) as conn:
            await conn.execute("SELECT 1")
            db_ok = True
    except Exception:
        pass
    return {
        "status": "ok" if db_ok else "degraded",
        "database": "connected" if db_ok else "error",
        "ai_model": config.AI_MODEL,
        "ledger_exists": Path(ledger_engine.LEDGER_PATH).exists(),
    }


@app.post("/api/chat", response_model=ChatResponse,
          summary="Send a message to Pacioli")
@limiter.limit(config.CHAT_RATE_LIMIT)
async def chat(request: Request, chat_req: ChatRequest,
               _auth: None = Depends(require_auth)):
    """
    Full lifecycle:
    1. Sanitise & store user message
    2. Build ledger context + history
    3. Call Gemini with retry on validation errors
    4. Validate & store proposal
    5. Return acknowledgement + proposal_id
    """
    if len(chat_req.message) > config.MAX_INPUT_LENGTH:
        raise HTTPException(400, f"Message exceeds {config.MAX_INPUT_LENGTH} characters")
    clean_msg = sanitize_input(chat_req.message)

    await db.insert_chat_message("user", clean_msg)
    history = await db.get_chat_history(limit=20)
    ledger_summary = ledger_engine.get_ledger_summary()
    chart_summary = ledger_engine.get_chart_text()

    last_error: str | None = None

    for attempt in range(config.MAX_RETRIES + 1):
        user_msg = clean_msg
        if last_error and attempt > 0:
            user_msg = (
                f"{clean_msg}\n\n"
                f"[SYSTEM FEEDBACK — fix these validation errors and retry:]\n"
                f"{last_error}\n\n"
                "Remember: write_cell needs cell_ref+new_value; "
                "insert_row needs row_index+values; debits must equal credits."
            )

        try:
            ai_response = await ai_client.call_model(
                ledger_summary=ledger_summary,
                chart_summary=chart_summary,
                chat_history=_trim_for_context(history[:-1]) if attempt == 0 else [],
                user_message=user_msg,
            )
        except Exception as e:
            logger.error("AI call failed: %s", e)
            err_msg = (
                "I'm having trouble connecting to the AI right now. "
                "Please try again in a moment."
            )
            await db.insert_chat_message("assistant", err_msg)
            return ChatResponse(assistant_message=err_msg)

        # Handle list responses (Gemini sometimes wraps in array)
        if isinstance(ai_response, list) and len(ai_response) > 0:
            ai_response = ai_response[0] if isinstance(ai_response[0], dict) else {"report": {"title": "Report"}}

        # ── Report (read-only, numbers computed from the ledger) ───
        if isinstance(ai_response, dict) and "report" in ai_response:
            report = ai_response["report"]
            # Compute the report from real ledger data — never trust AI numbers (item 4.1).
            try:
                computed = report_engine.generate(report.get("title", ""), clean_msg)
            except Exception as e:
                logger.error("Report computation failed: %s", e)
                computed = None

            if computed:
                report_text = report_engine.render_markdown(computed)
            else:
                # Unknown report type — fall back to the AI's structure.
                report_text = _render_ai_report(report)

            await db.insert_chat_message("assistant", report_text)
            return ChatResponse(assistant_message=report_text)

        # ── Transaction proposal ──────────────────────────────────
        if not isinstance(ai_response, dict):
            msg = "The AI response was not in the expected format. Please try rephrasing."
            await db.insert_chat_message("assistant", msg)
            return ChatResponse(assistant_message=msg)
        proposal_data = ai_response.get("proposal")
        if not isinstance(proposal_data, dict):
            if proposal_data is not None:
                msg = "The AI response was not in the expected format. Please try rephrasing."
            else:
                msg = "I could not form a clear proposal. Could you rephrase your request?"
            await db.insert_chat_message("assistant", msg)
            return ChatResponse(assistant_message=msg)

        try:
            validated = Proposal(**proposal_data)
            last_error = None
            break
        except Exception as e:
            last_error = str(e)
            if attempt == config.MAX_RETRIES:
                msg = (
                    "I'm having trouble generating a valid proposal. "
                    "Could you rephrase your request more simply?\n\n"
                    f"_(Technical detail: {e})_"
                )
                await db.insert_chat_message("assistant", msg)
                return ChatResponse(assistant_message=msg)
            continue

    # ── Equation check ────────────────────────────────────────────
    eq = validated.accounting_equation_check
    if not eq.balance_confirmed:
        msg = (
            "I could not confirm the accounting equation balances. "
            "I won't submit this proposal. Please verify your instruction."
        )
        await db.insert_chat_message("assistant", msg)
        return ChatResponse(assistant_message=msg)

    # ── Fiscal-period check (non-blocking warnings, item 1.3) ──────
    warnings = fiscal_warnings(validated.actions)

    # ── Store proposal ────────────────────────────────────────────
    proposal_id = await db.create_proposal(
        user_message=clean_msg,
        ai_reasoning=validated.justification,
        actions=[a.model_dump() for a in validated.actions],
        validation_notes=warnings,
    )

    warn_text = ("\n\n⚠️ " + " ".join(warnings)) if warnings else ""
    ack = (
        f"I've prepared a proposed edit for your review. (Proposal #{proposal_id})\n\n"
        f"**Summary:** {validated.summary}\n"
        f"**Reasoning:** {validated.justification}{warn_text}\n\n"
        "Please review the changes in the right panel and approve or reject."
    )
    await db.insert_chat_message("assistant", ack, proposal_id=proposal_id)

    return ChatResponse(
        assistant_message=ack,
        proposal_id=proposal_id,
        proposal_summary=validated.summary,
    )


def _parse_notes(raw) -> list[str]:
    if not raw:
        return []
    try:
        val = json.loads(raw)
        return val if isinstance(val, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


@app.get("/api/proposals", summary="List recent proposals")
async def list_proposals(limit: int = 10, offset: int = 0,
                         _auth: None = Depends(require_auth)):
    proposals, total = await db.get_proposals(limit=min(limit, 50), offset=max(0, offset))
    return {"proposals": proposals, "total": total}


@app.get("/api/proposals/{proposal_id}", response_model=ProposalDetail,
         summary="Fetch a proposal for preview")
async def get_proposal(proposal_id: int,
                       _auth: None = Depends(require_auth)):
    proposal = await db.get_proposal(proposal_id)
    if not proposal:
        raise HTTPException(404, "Proposal not found")

    preview_actions = []
    for act in proposal["actions"]:
        old_val = None
        if act.get("cell_ref") and act["operation"] in ("write_cell", "write_formula"):
            wb = ledger_engine.get_workbook(data_only=True)
            try:
                if act["sheet"] in wb.sheetnames:
                    old_val = wb[act["sheet"]][act["cell_ref"]].value
            except Exception:
                old_val = "N/A"
            finally:
                wb.close()
        preview_actions.append({
            **act,
            "old_value_display": act.get("old_value", old_val),
            "new_value_display": act.get("new_value", act.get("formula", "")),
        })

    # Build highlight map: sheet → list of cell refs that will change
    highlight_cells: dict[str, list[str]] = {}
    for act in proposal["actions"]:
        sheet = act.get("sheet", "")
        cell = act.get("cell_ref")
        if cell:
            highlight_cells.setdefault(sheet, []).append(cell)

    return ProposalDetail(
        id=proposal["id"],
        status=proposal["status"],
        justification=proposal.get("ai_reasoning", ""),
        actions=preview_actions,
        user_message=proposal["user_message"],
        created_at=proposal["created_at"],
        validation_notes=_parse_notes(proposal.get("validation_notes")),
        highlight_cells=highlight_cells,
    )


@app.post("/api/proposals/{proposal_id}/approve", response_model=ApprovalResponse,
          summary="Approve and execute a proposal")
async def approve_proposal(proposal_id: int,
                           _auth: None = Depends(require_auth)):
    proposal = await db.get_proposal(proposal_id)
    if not proposal:
        raise HTTPException(404, "Proposal not found")

    # Atomic status transition — prevents double-approval race condition
    approved = await db.approve_proposal_atomic(proposal_id)
    if not approved:
        raise HTTPException(
            409, f"Proposal is already '{proposal['status']}', not 'pending'."
        )

    try:
        # The snapshot is captured inside the lock by execute_actions, so it is a
        # consistent pre-execution state for undo.
        snapshot, change_log = ledger_engine.execute_actions(proposal["actions"])
    except ledger_engine.LockTimeout:
        # Nothing was written (the atomic save never ran); let the user retry.
        await db.reset_proposal_pending(proposal_id)
        raise HTTPException(
            409, "The ledger is busy right now. Please try again in a moment."
        )
    except Exception as e:
        logger.error("Execution failed for proposal %d: %s", proposal_id, e)
        await db.update_proposal_status(proposal_id, "failed", str(e))
        await db.insert_chat_message(
            "assistant",
            f"Proposal #{proposal_id} execution failed: {e}",
            proposal_id=proposal_id,
        )
        return ApprovalResponse(success=False, message=f"Execution failed: {e}")

    await db.save_snapshot(proposal_id, snapshot)
    # Persist an audit trail of what was executed (item 2.9)
    for idx, act in enumerate(proposal["actions"]):
        await db.insert_audit_log(
            proposal_id=proposal_id,
            action_index=idx,
            sheet=act.get("sheet", ""),
            cell_ref=act.get("cell_ref"),
            old_value=act.get("old_value"),
            new_value=act.get("new_value") or act.get("formula")
            or act.get("values"),
        )
    await db.insert_chat_message(
        "assistant",
        f"Proposal #{proposal_id} approved and executed. "
        f"Changes: {'; '.join(change_log)}",
        proposal_id=proposal_id,
    )
    return ApprovalResponse(
        success=True,
        message="Proposal executed successfully.",
        change_log=change_log,
    )


@app.post("/api/proposals/{proposal_id}/reject", response_model=ApprovalResponse,
          summary="Reject a proposal")
async def reject_proposal(proposal_id: int,
                          _auth: None = Depends(require_auth)):
    proposal = await db.get_proposal(proposal_id)
    if not proposal:
        raise HTTPException(404, "Proposal not found")
    if proposal["status"] != "pending":
        raise HTTPException(409, f"Proposal is '{proposal['status']}', not 'pending'.")

    await db.update_proposal_status(proposal_id, "rejected")
    await db.insert_chat_message(
        "assistant",
        f"Proposal #{proposal_id} rejected. No edits were made to the ledger.",
        proposal_id=proposal_id,
    )
    return ApprovalResponse(success=True, message="Proposal rejected.")


@app.post("/api/proposals/{proposal_id}/restore", response_model=ApprovalResponse,
          summary="Restore the ledger to its state before this proposal (undo)")
async def restore_proposal(proposal_id: int,
                           _auth: None = Depends(require_auth)):
    """Roll the ledger back to the snapshot taken before this proposal executed."""
    proposal = await db.get_proposal(proposal_id)
    if not proposal:
        raise HTTPException(404, "Proposal not found")

    snapshot = await db.get_snapshot(proposal_id)
    if snapshot is None:
        raise HTTPException(404, "No snapshot available for this proposal.")

    try:
        ledger_engine.restore_snapshot(snapshot)
    except Exception as e:
        logger.error("Restore failed for proposal %d: %s", proposal_id, e)
        raise HTTPException(500, f"Restore failed: {e}")

    await db.insert_audit_log(
        proposal_id=proposal_id, action_index=-1, sheet="*",
        cell_ref=None, old_value="executed", new_value="restored",
    )
    await db.insert_chat_message(
        "assistant",
        f"Ledger restored to the state before proposal #{proposal_id}.",
        proposal_id=proposal_id,
    )
    return ApprovalResponse(
        success=True,
        message=f"Ledger restored to before proposal #{proposal_id}.",
    )


@app.get("/api/ledger/preview", summary="Read-only preview of a ledger sheet")
async def ledger_preview(sheet: str = "GeneralLedger", limit: int = 50,
                         _auth: None = Depends(require_auth)):
    """Return the header row + first N data rows of a sheet as JSON (item 1.7)."""
    try:
        wb = ledger_engine.get_workbook(data_only=True)
    except Exception as e:
        raise HTTPException(500, f"Could not open ledger: {e}")
    try:
        if sheet not in wb.sheetnames:
            raise HTTPException(404, f"Sheet '{sheet}' not found.")
        ws = wb[sheet]
        max_col = ws.max_column or 1
        headers = [str(ws.cell(1, c).value or "") for c in range(1, max_col + 1)]
        rows = []
        for r in range(2, min(ws.max_row, 1 + max(1, limit)) + 1):
            rows.append([
                ws.cell(r, c).value for c in range(1, max_col + 1)
            ])
        return {"sheet": sheet, "sheets": wb.sheetnames,
                "headers": headers, "rows": rows}
    finally:
        wb.close()


@app.get("/api/chat/history", summary="Return full chat history")
async def get_history(_auth: None = Depends(require_auth)):
    history = await db.get_chat_history(limit=100)
    return {"messages": history}


@app.get("/api/audit", summary="Return audit log entries")
async def get_audit(limit: int = 50,
                    _auth: None = Depends(require_auth)):
    async with aiosqlite.connect(db.DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute(
            "SELECT * FROM audit_log ORDER BY id DESC LIMIT ?", (min(limit, 200),)
        )
        rows = await cursor.fetchall()
    return {"entries": [dict(r) for r in rows]}


@app.get("/api/ledger/download", summary="Download the current ledger file")
async def download_ledger(_auth: None = Depends(require_auth)):
    if not Path(ledger_engine.LEDGER_PATH).exists():
        raise HTTPException(404, "Ledger file not found.")
    return FileResponse(
        ledger_engine.LEDGER_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="ledger.xlsx",
    )


def _parse_transaction_row(new_value: str) -> dict:
    """Parse a stringified insert_row values list into named fields.

    GeneralLedger columns: Date, Ref, Description, Account, AccountName, Debit, Credit, Balance
    """
    try:
        vals = json.loads(new_value.replace("'", '"'))
        if not isinstance(vals, list):
            return {}
        return {
            "date": str(vals[0]) if len(vals) > 0 else "",
            "ref": str(vals[1]) if len(vals) > 1 else "",
            "description": str(vals[2]) if len(vals) > 2 else "",
            "account": str(vals[3]) if len(vals) > 3 else "",
            "account_name": str(vals[4]) if len(vals) > 4 else "",
            "debit": vals[5] if len(vals) > 5 else 0,
            "credit": vals[6] if len(vals) > 6 else 0,
        }
    except (json.JSONDecodeError, TypeError, IndexError):
        return {}


@app.get("/api/transactions", summary="List executed transactions from audit log")
async def get_transactions(limit: int = 50, search: str = "",
                           _auth: None = Depends(require_auth)):
    async with aiosqlite.connect(db.DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        query = """
            SELECT a.*, p.user_message, p.created_at as proposal_date
            FROM audit_log a
            LEFT JOIN proposals p ON a.proposal_id = p.id
            WHERE a.action_index >= 0
        """
        params = []
        if search:
            query += " AND (a.new_value LIKE ? OR p.user_message LIKE ?)"
            params.extend([f"%{search}%", f"%{search}%"])
        query += " ORDER BY a.id DESC LIMIT ?"
        params.append(min(limit, 200))
        cursor = await conn.execute(query, params)
        rows = await cursor.fetchall()

    transactions = []
    for r in rows:
        d = dict(r)
        parsed = _parse_transaction_row(d.get("new_value") or "")
        d["date"] = parsed.get("date", "")
        d["account"] = parsed.get("account_name", "") or parsed.get("account", "")
        d["debit"] = parsed.get("debit", 0)
        d["credit"] = parsed.get("credit", 0)
        d["description"] = parsed.get("description", "") or d.get("user_message", "")
        transactions.append(d)
    return {"transactions": transactions}


@app.get("/api/reports/{report_type}/csv", summary="Download report as CSV")
async def download_report_csv(report_type: str,
                              _auth: None = Depends(require_auth)):
    import csv
    import io
    from fastapi.responses import StreamingResponse

    # Map URL slug to internal name
    type_map = {
        "trial-balance": "trial_balance",
        "income-statement": "income_statement",
        "balance-sheet": "balance_sheet",
    }
    internal_type = type_map.get(report_type, report_type)

    if internal_type == "trial_balance":
        report = report_engine.trial_balance()
        rows = report.get("rows", [])
        headers = ["Account", "Name", "Debit", "Credit"]
        data = [[r["account"], r["name"], r["debit"], r["credit"]] for r in rows]
        data.append(["", "TOTAL", report["total_debit"], report["total_credit"]])
    elif internal_type == "income_statement":
        report = report_engine.income_statement()
        rows = report.get("revenue", []) + report.get("expenses", [])
        headers = ["Account", "Name", "Amount"]
        data = [[r["account"], r["name"], r["amount"]] for r in rows]
        data.append(["", "Net Income", report["net_income"]])
    elif internal_type == "balance_sheet":
        report = report_engine.balance_sheet()
        sections = report.get("assets", []) + report.get("liabilities", []) + report.get("equity", [])
        headers = ["Account", "Name", "Amount"]
        data = [[r["account"], r["name"], r["amount"]] for r in sections]
        data.append(["", "Total Assets", report["total_assets"]])
        data.append(["", "Total Liabilities + Equity", report["total_liabilities"] + report["total_equity"]])
    else:
        raise HTTPException(404, f"Unknown report type: {report_type}")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(data)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={report_type}.csv"},
    )


@app.get("/api/reports/{report_type}/xlsx", summary="Download report as XLSX")
async def download_report_xlsx(report_type: str,
                               _auth: None = Depends(require_auth)):
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io

    type_map = {
        "trial-balance": "trial_balance",
        "income-statement": "income_statement",
        "balance-sheet": "balance_sheet",
    }
    internal_type = type_map.get(report_type, report_type)

    wb = Workbook()
    ws = wb.active

    if internal_type == "trial_balance":
        report = report_engine.trial_balance()
        ws.title = "Trial Balance"
        ws.append(["Account", "Name", "Debit", "Credit"])
        for r in report.get("rows", []):
            ws.append([r["account"], r["name"], r["debit"], r["credit"]])
        ws.append(["", "TOTAL", report["total_debit"], report["total_credit"]])
    elif internal_type == "income_statement":
        report = report_engine.income_statement()
        ws.title = "Income Statement"
        ws.append(["Account", "Name", "Amount"])
        for r in report.get("revenue", []):
            ws.append([r["account"], r["name"], r["amount"]])
        ws.append(["", "Total Revenue", report["total_revenue"]])
        ws.append([])
        for r in report.get("expenses", []):
            ws.append([r["account"], r["name"], r["amount"]])
        ws.append(["", "Total Expenses", report["total_expenses"]])
        ws.append(["", "Net Income", report["net_income"]])
    elif internal_type == "balance_sheet":
        report = report_engine.balance_sheet()
        ws.title = "Balance Sheet"
        ws.append(["Account", "Name", "Amount"])
        ws.append(["--- Assets ---", "", ""])
        for r in report.get("assets", []):
            ws.append([r["account"], r["name"], r["amount"]])
        ws.append(["", "Total Assets", report["total_assets"]])
        ws.append([])
        ws.append(["--- Liabilities ---", "", ""])
        for r in report.get("liabilities", []):
            ws.append([r["account"], r["name"], r["amount"]])
        ws.append(["", "Total Liabilities", report["total_liabilities"]])
        ws.append([])
        ws.append(["--- Equity ---", "", ""])
        for r in report.get("equity", []):
            ws.append([r["account"], r["name"], r["amount"]])
        ws.append(["", "Net Income", report["net_income"]])
        ws.append(["", "Total Equity", report["total_equity"]])
        ws.append(["", "Total L+E", report["total_liabilities"] + report["total_equity"]])
    else:
        raise HTTPException(404, f"Unknown report type: {report_type}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report_type}.xlsx"},
    )


@app.get("/api/transactions/csv", summary="Download transactions as CSV")
async def download_transactions_csv(_auth: None = Depends(require_auth)):
    import csv
    import io
    from fastapi.responses import StreamingResponse

    async with aiosqlite.connect(db.DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute("""
            SELECT a.*, p.user_message, p.created_at as proposal_date
            FROM audit_log a
            LEFT JOIN proposals p ON a.proposal_id = p.id
            WHERE a.action_index >= 0
            ORDER BY a.id DESC
        """)
        rows = await cursor.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Description", "Account", "Debit", "Credit", "Proposal"])
    for r in rows:
        parsed = _parse_transaction_row(r["new_value"] or "")
        writer.writerow([
            parsed.get("date", ""),
            parsed.get("description", "") or r["user_message"],
            parsed.get("account_name", "") or parsed.get("account", ""),
            parsed.get("debit", 0),
            parsed.get("credit", 0),
            r["proposal_id"],
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=transactions.csv"},
    )


# ── Personal Bookkeeping Endpoints ────────────────────────────

@app.get("/api/accounts", summary="List all accounts")
async def list_accounts(_auth: None = Depends(require_auth)):
    accounts = await db.get_accounts()
    # Add computed balance to each account
    for acc in accounts:
        acc["balance"] = await db.get_account_balance(acc["id"])
    return {"accounts": accounts}


@app.post("/api/accounts", summary="Create an account")
async def create_account_endpoint(name: str, acc_type: str, currency: str = "USD",
                                   _auth: None = Depends(require_auth)):
    acc_id = await db.create_account(name, acc_type, currency)
    return {"id": acc_id, "name": name, "type": acc_type}


@app.put("/api/accounts/{account_id}", summary="Update an account")
async def update_account_endpoint(account_id: int, name: str = None, acc_type: str = None,
                                   _auth: None = Depends(require_auth)):
    await db.update_account(account_id, name, acc_type)
    return {"success": True}


@app.delete("/api/accounts/{account_id}", summary="Deactivate an account")
async def delete_account_endpoint(account_id: int, _auth: None = Depends(require_auth)):
    await db.delete_account(account_id)
    return {"success": True}


@app.get("/api/categories", summary="List all categories")
async def list_categories(_auth: None = Depends(require_auth)):
    cats = await db.get_categories()
    return {"categories": cats}


@app.post("/api/categories", summary="Create a category")
async def create_category_endpoint(name: str, parent_id: int = None, icon: str = "",
                                    _auth: None = Depends(require_auth)):
    cat_id = await db.create_category(name, parent_id, icon)
    return {"id": cat_id, "name": name}


@app.put("/api/categories/{category_id}", summary="Update a category")
async def update_category_endpoint(category_id: int, name: str = None, icon: str = None,
                                    _auth: None = Depends(require_auth)):
    await db.update_category(category_id, name, icon)
    return {"success": True}


@app.get("/api/transactions/list", summary="List transactions with filters")
async def list_transactions(limit: int = 50, offset: int = 0,
                            account_id: int = None, category_id: int = None,
                            _auth: None = Depends(require_auth)):
    txs = await db.get_transactions(limit, offset, account_id, category_id)
    return {"transactions": txs}


@app.post("/api/transactions/{tx_id}/categorize", summary="Assign category to transaction")
async def categorize_transaction(tx_id: int, category_id: int,
                                  _auth: None = Depends(require_auth)):
    await db.update_transaction_category(tx_id, category_id)
    # Learn from this categorization
    txs = await db.get_transactions(limit=1)
    for tx in txs:
        if tx["id"] == tx_id and tx.get("description"):
            await db.learn_rule(tx["description"], category_id)
            break
    return {"success": True}


@app.get("/api/rules", summary="List categorization rules")
async def list_rules(_auth: None = Depends(require_auth)):
    rules = await db.get_rules()
    return {"rules": rules}


@app.post("/api/rules", summary="Create a categorization rule")
async def create_rule_endpoint(pattern: str, category_id: int, account_id: int = None,
                                _auth: None = Depends(require_auth)):
    rule_id = await db.create_rule(pattern, category_id, account_id)
    return {"id": rule_id}


@app.delete("/api/rules/{rule_id}", summary="Delete a rule")
async def delete_rule_endpoint(rule_id: int, _auth: None = Depends(require_auth)):
    await db.delete_rule(rule_id)
    return {"success": True}


# ── Recurring Transactions ────────────────────────────────────

@app.get("/api/recurring", summary="List recurring transactions")
async def list_recurring(_auth: None = Depends(require_auth)):
    items = await db.get_recurring()
    return {"recurring": items}


@app.post("/api/recurring", summary="Create recurring transaction")
async def create_recurring_endpoint(account_id: int, category_id: int, description: str,
                                     amount: float, tx_type: str, frequency: str,
                                     next_date: str,
                                     _auth: None = Depends(require_auth)):
    rid = await db.create_recurring(account_id, category_id, description, amount, tx_type, frequency, next_date)
    return {"id": rid}


@app.put("/api/recurring/{recurring_id}", summary="Update recurring transaction")
async def update_recurring_endpoint(recurring_id: int, description: str = None,
                                     amount: float = None, frequency: str = None,
                                     next_date: str = None, is_active: int = None,
                                     category_id: int = None, account_id: int = None,
                                     _auth: None = Depends(require_auth)):
    kwargs = {k: v for k, v in {"description": description, "amount": amount,
              "frequency": frequency, "next_date": next_date, "is_active": is_active,
              "category_id": category_id, "account_id": account_id}.items() if v is not None}
    await db.update_recurring(recurring_id, **kwargs)
    return {"success": True}


@app.delete("/api/recurring/{recurring_id}", summary="Deactivate recurring transaction")
async def delete_recurring_endpoint(recurring_id: int, _auth: None = Depends(require_auth)):
    await db.delete_recurring(recurring_id)
    return {"success": True}


@app.post("/api/recurring/{recurring_id}/execute", summary="Execute a recurring transaction now")
async def execute_recurring_endpoint(recurring_id: int, _auth: None = Depends(require_auth)):
    items = await db.get_recurring()
    item = None
    for r in items:
        if r["id"] == recurring_id:
            item = r
            break
    if not item:
        raise HTTPException(404, "Recurring transaction not found")
    today = datetime.utcnow().strftime("%Y-%m-%d")
    await db.insert_transaction(
        proposal_id=None, account_id=item["account_id"], category_id=item["category_id"],
        date=today, description=item["description"], amount=item["amount"],
        tx_type=item["type"],
    )
    await db.advance_recurring(recurring_id, item["frequency"], today)
    return {"success": True, "message": f"Executed: {item['description']}"}


# ── Budgets ───────────────────────────────────────────────────

@app.get("/api/budgets", summary="List budgets for a month")
async def list_budgets(year: int = None, month: int = None,
                       _auth: None = Depends(require_auth)):
    if not year or not month:
        now = datetime.utcnow()
        year, month = now.year, now.month
    budgets = await db.get_budget_status(year, month)
    return {"budgets": budgets, "year": year, "month": month}


@app.post("/api/budgets", summary="Set budget for a category")
async def set_budget_endpoint(category_id: int, amount: float,
                               year: int = None, month: int = None,
                               _auth: None = Depends(require_auth)):
    if not year or not month:
        now = datetime.utcnow()
        year, month = now.year, now.month
    bid = await db.set_budget(category_id, amount, year, month)
    return {"id": bid}


@app.delete("/api/budgets/{budget_id}", summary="Remove a budget")
async def delete_budget_endpoint(budget_id: int, _auth: None = Depends(require_auth)):
    await db.delete_budget(budget_id)
    return {"success": True}


# ── Transfers ─────────────────────────────────────────────────

@app.post("/api/transfers", summary="Create a transfer between accounts")
async def create_transfer_endpoint(from_account_id: int, to_account_id: int,
                                    amount: float, description: str,
                                    date: str = None,
                                    _auth: None = Depends(require_auth)):
    if from_account_id == to_account_id:
        raise HTTPException(400, "Cannot transfer to the same account")
    if amount <= 0:
        raise HTTPException(400, "Amount must be positive")
    if not date:
        date = datetime.utcnow().strftime("%Y-%m-%d")
    pair_id = await db.create_transfer(from_account_id, to_account_id, amount, description, date)
    return {"transfer_pair_id": pair_id, "success": True}


# ── Export & Backup ───────────────────────────────────────────

@app.get("/api/export/csv", summary="Export all transactions as CSV")
async def export_transactions_csv(_auth: None = Depends(require_auth)):
    import csv
    import io
    from fastapi.responses import StreamingResponse

    txs = await db.get_transactions(limit=10000)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Description", "Account", "Category", "Amount", "Type", "Reference"])
    for tx in txs:
        writer.writerow([
            tx.get("date", ""), tx.get("description", ""),
            tx.get("account_name", ""), tx.get("category_name", ""),
            tx.get("amount", 0), tx.get("type", ""),
            tx.get("reference", ""),
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=transactions.csv"},
    )


@app.get("/api/export/json", summary="Export all data as JSON")
async def export_data_json(_auth: None = Depends(require_auth)):
    from fastapi.responses import JSONResponse
    accounts = await db.get_accounts()
    for a in accounts:
        a["balance"] = await db.get_account_balance(a["id"])
    categories = await db.get_categories()
    transactions = await db.get_transactions(limit=10000)
    rules = await db.get_rules()
    recurring = await db.get_recurring()
    return JSONResponse({
        "accounts": accounts,
        "categories": categories,
        "transactions": transactions,
        "rules": rules,
        "recurring": recurring,
    })


@app.get("/api/backups", summary="List available backups")
async def list_backups(_auth: None = Depends(require_auth)):
    import glob as glob_mod
    backups = sorted(glob_mod.glob("data/ledger_backup_*.xlsx"), reverse=True)
    return {"backups": [Path(b).name for b in backups[:config.BACKUP_MAX_COUNT]]}


# ── Tax-ready Reports ────────────────────────────────────────

@app.get("/api/tax/summary", summary="Annual tax summary")
async def tax_summary(year: int = None, _auth: None = Depends(require_auth)):
    if not year:
        year = datetime.utcnow().year
    summary = await db.get_tax_summary(year)
    return summary


@app.get("/api/tax/transactions", summary="List transactions by tax tag")
async def tax_transactions(year: int = None, tag: str = None,
                           _auth: None = Depends(require_auth)):
    if not year:
        year = datetime.utcnow().year
    txs = await db.get_tagged_transactions(year, tag)
    return {"transactions": txs}


@app.post("/api/tax/tag", summary="Tag a transaction for tax purposes")
async def tag_transaction_endpoint(transaction_id: int, tag: str, notes: str = "",
                                    _auth: None = Depends(require_auth)):
    await db.tag_transaction(transaction_id, tag, notes)
    return {"success": True}


@app.delete("/api/tax/tag/{transaction_id}", summary="Remove tax tag")
async def untag_transaction_endpoint(transaction_id: int,
                                      _auth: None = Depends(require_auth)):
    await db.untag_transaction(transaction_id)
    return {"success": True}


@app.post("/api/tax/auto-tag", summary="Auto-tag transactions by category")
async def auto_tag_endpoint(year: int = None, _auth: None = Depends(require_auth)):
    if not year:
        year = datetime.utcnow().year
    count = await db.auto_tag_transactions(year, config.TAX_CATEGORY_DEFAULTS)
    return {"success": True, "tagged": count}


@app.get("/api/tax/export", summary="Export tax summary as CSV")
async def tax_export_csv(year: int = None, _auth: None = Depends(require_auth)):
    import csv
    import io
    from fastapi.responses import StreamingResponse

    if not year:
        year = datetime.utcnow().year
    txs = await db.get_tagged_transactions(year)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Description", "Account", "Category", "Amount", "Type", "Tax Tag", "Notes"])
    for tx in txs:
        writer.writerow([
            tx.get("date", ""), tx.get("description", ""),
            tx.get("account_name", ""), tx.get("category_name", ""),
            tx.get("amount", 0), tx.get("type", ""),
            tx.get("tag", "untagged"), tx.get("notes", ""),
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=tax-report-{year}.csv"},
    )


# ── Bank CSV Import ──────────────────────────────────────────

@app.get("/api/csv/profiles", summary="List saved CSV import profiles")
async def list_csv_profiles(_auth: None = Depends(require_auth)):
    profiles = await db.get_csv_profiles()
    return {"profiles": profiles}


@app.post("/api/csv/profiles", summary="Save a CSV import profile")
async def create_csv_profile_endpoint(name: str, delimiter: str = ",",
                                       date_col: int = 0, desc_col: int = 1,
                                       amount_col: int = 2, has_header: int = 1,
                                       date_format: str = "YYYY-MM-DD",
                                       amount_positive: str = "positive",
                                       _auth: None = Depends(require_auth)):
    pid = await db.create_csv_profile(name, delimiter, date_col, desc_col,
                                       amount_col, has_header, date_format, amount_positive)
    return {"id": pid}


@app.delete("/api/csv/profiles/{profile_id}", summary="Delete a CSV profile")
async def delete_csv_profile_endpoint(profile_id: int, _auth: None = Depends(require_auth)):
    await db.delete_csv_profile(profile_id)
    return {"success": True}


@app.post("/api/csv/preview", summary="Parse CSV and preview first rows")
async def csv_preview(csv_text: str, delimiter: str = ",",
                      _auth: None = Depends(require_auth)):
    import csv
    import io
    reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
    rows = []
    headers = None
    for i, row in enumerate(reader):
        if i == 0:
            headers = row
        elif i <= 5:
            rows.append(row)
        else:
            break
    return {"headers": headers or [], "preview": rows, "total_rows": csv_text.strip().count("\n") + 1}


@app.post("/api/csv/import", summary="Import CSV transactions")
async def csv_import(csv_text: str, delimiter: str = ",",
                     date_col: int = 0, desc_col: int = 1, amount_col: int = 2,
                     account_id: int = None, category_id: int = None,
                     date_format: str = "YYYY-MM-DD",
                     amount_positive: str = "positive",
                     profile_name: str = None,
                     _auth: None = Depends(require_auth)):
    import csv
    import io
    from datetime import datetime as _dt

    # Save profile if requested
    if profile_name:
        await db.create_csv_profile(profile_name, delimiter, date_col, desc_col,
                                     amount_col, 1, date_format, amount_positive)

    reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return {"success": False, "message": "Empty CSV"}

    # Skip header if first row looks like headers
    start = 1 if rows[0][0].isalpha() else 0

    # Get default account if not specified
    if not account_id:
        accounts = await db.get_accounts()
        account_id = accounts[0]["id"] if accounts else None
    if not account_id:
        return {"success": False, "message": "No account specified and no accounts exist"}

    imported = 0
    skipped = 0
    today = _dt.utcnow().strftime("%Y-%m-%d")

    for row in rows[start:]:
        try:
            # Parse date
            date_str = row[date_col].strip() if date_col < len(row) else today
            if date_format == "MM/DD/YYYY" and "/" in date_str:
                parts = date_str.split("/")
                date_str = f"{parts[2]}-{parts[0]:0>2}-{parts[1]:0>2}"
            elif date_format == "DD/MM/YYYY" and "/" in date_str:
                parts = date_str.split("/")
                date_str = f"{parts[2]}-{parts[1]:0>2}-{parts[0]:0>2}"

            # Parse description
            desc = row[desc_col].strip() if desc_col < len(row) else ""

            # Parse amount
            amount_str = row[amount_col].strip().replace("$", "").replace(",", "") if amount_col < len(row) else "0"
            amount = float(amount_str)
            if amount_positive == "absolute":
                # Heuristic: negative if description contains certain keywords
                neg_words = ["payment", "purchase", "debit", "fee", "charge", "withdrawal"]
                if any(w in desc.lower() for w in neg_words):
                    amount = -abs(amount)
                elif amount > 0:
                    amount = -abs(amount)  # Default to expense for absolute mode

            # Determine type
            tx_type = "income" if amount > 0 else "expense"

            # Auto-categorize
            cat_id = category_id
            if not cat_id:
                rule = await db.match_rule(desc)
                if rule:
                    cat_id = rule["category_id"]

            await db.insert_transaction(
                proposal_id=None, account_id=account_id, category_id=cat_id,
                date=date_str, description=desc, amount=amount, tx_type=tx_type,
            )
            imported += 1
        except (ValueError, IndexError):
            skipped += 1

    return {"success": True, "imported": imported, "skipped": skipped}


# ── Updated Dashboard with monthly trend ─────────────────────

@app.get("/api/dashboard", summary="Dashboard data")
async def get_dashboard(_auth: None = Depends(require_auth)):
    now = datetime.utcnow()
    year, month = now.year, now.month

    accounts = await db.get_accounts()
    for acc in accounts:
        acc["balance"] = await db.get_account_balance(acc["id"])

    summary = await db.get_monthly_summary(year, month)
    breakdown = await db.get_category_breakdown(year, month)
    recent = await db.get_transactions(limit=10)

    total_expenses = summary["expenses"] or 1
    for item in breakdown:
        item["pct"] = round((item["total"] / total_expenses) * 100, 1)

    # Monthly trend (last 6 months)
    monthly_trend = []
    for i in range(5, -1, -1):
        m = month - i
        y = year
        while m <= 0:
            m += 12
            y -= 1
        ms = await db.get_monthly_summary(y, m)
        monthly_trend.append({
            "month": f"{y}-{m:02d}",
            "label": datetime(y, m, 1).strftime("%b"),
            "income": ms["income"],
            "expenses": ms["expenses"],
        })

    # Budget status
    budget_status = await db.get_budget_status(year, month)

    return {
        "accounts": accounts,
        "summary": {
            "total_income": summary["income"],
            "total_expenses": summary["expenses"],
            "net": summary["net"],
            "period": f"{now.strftime('%B %Y')}",
        },
        "by_category": breakdown,
        "recent_transactions": recent,
        "monthly_trend": monthly_trend,
        "budgets": budget_status,
    }


@app.get("/", response_class=HTMLResponse)
async def serve_dashboard():
    html_path = Path("static/index.html")
    if html_path.exists():
        return HTMLResponse(html_path.read_text())
    return HTMLResponse("<h1>Pacioli</h1><p>Place static/index.html</p>")
