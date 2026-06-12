"""Safe read/write operations on the Excel ledger with snapshot support."""

import contextlib
import logging
import os
import time
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from config import LEDGER_PATH

logger = logging.getLogger(__name__)

# ── File locking ──────────────────────────────────────────────────
try:
    from filelock import FileLock as _FileLock
    from filelock import Timeout as LockTimeout

    def _get_lock():
        return _FileLock(LEDGER_PATH + ".lock", timeout=10)
except ImportError:
    class LockTimeout(Exception):
        """Fallback when filelock isn't installed (locking degrades to a no-op)."""

    def _get_lock():
        return contextlib.nullcontext()

# ── Caches ────────────────────────────────────────────────────────
_summary_cache: dict = {"data": None, "ts": 0.0}
_column_cache: dict = {}


def _invalidate_caches():
    _summary_cache["data"] = None
    _summary_cache["ts"] = 0.0
    _column_cache.clear()


# ── Workbook helpers ──────────────────────────────────────────────

def get_workbook(data_only: bool = False):
    return load_workbook(LEDGER_PATH, data_only=data_only)


def take_snapshot() -> bytes:
    with open(LEDGER_PATH, "rb") as f:
        return f.read()


def _atomic_replace_with_bytes(data: bytes):
    """Write ``data`` to a temp file beside the ledger, then atomically swap it in.

    ``os.replace`` is atomic on a single filesystem, so a crash mid-write leaves
    either the complete old file or the complete new one — never a torn .xlsx.
    """
    tmp_path = LEDGER_PATH + ".tmp"
    with open(tmp_path, "wb") as f:
        f.write(data)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp_path, LEDGER_PATH)


def get_chart_text() -> str:
    """Render the live Chart of Accounts as text for the system prompt.

    Sourced from the workbook (not hardcoded in the prompt) so the model can
    never propose an account the ledger will subsequently reject.
    """
    wb = get_workbook(data_only=True)
    try:
        if "ChartOfAccounts" not in wb.sheetnames:
            return "(no chart of accounts found)"
        cols = get_column_indices("ChartOfAccounts", wb)
        c_acct = cols.get("Account", 1)
        c_name = cols.get("AccountName", 2)
        c_type = cols.get("Type", 3)
        ws = wb["ChartOfAccounts"]
        lines = []
        for r in range(2, ws.max_row + 1):
            num = ws.cell(r, c_acct).value
            if num is None or str(num).strip() == "":
                continue
            name = str(ws.cell(r, c_name).value or "").strip()
            typ = str(ws.cell(r, c_type).value or "").strip()
            lines.append(f"- {str(num).strip()} {name} ({typ})")
        return "\n".join(lines) if lines else "(chart of accounts is empty)"
    finally:
        wb.close()


def get_column_indices(sheet_name: str, wb=None) -> dict[str, int]:
    """Map header names → 1-based column index for a sheet (cached)."""
    if sheet_name in _column_cache:
        return _column_cache[sheet_name]
    owns_wb = wb is None
    if owns_wb:
        wb = get_workbook(data_only=True)
    try:
        ws = wb[sheet_name]
        indices: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h:
                indices[str(h).strip()] = c
        _column_cache[sheet_name] = indices
        return indices
    finally:
        if owns_wb:
            wb.close()


def get_valid_accounts(wb=None) -> set[str]:
    """Return valid account numbers from the ChartOfAccounts sheet."""
    owns_wb = wb is None
    if owns_wb:
        wb = get_workbook(data_only=True)
    try:
        if "ChartOfAccounts" not in wb.sheetnames:
            return set()
        ws = wb["ChartOfAccounts"]
        return {
            str(row[0]).strip()
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row and row[0] is not None
        }
    finally:
        if owns_wb:
            wb.close()


def _validate_cell_bounds(ws, cell_ref: str):
    """Raise ValueError if the cell is unreasonably far from existing data."""
    try:
        _, row = coordinate_from_string(cell_ref)
        max_row = ws.max_row or 1
        if row > max_row + 10:
            raise ValueError(
                f"Cell {cell_ref} (row {row}) is too far from existing data "
                f"(sheet max row: {max_row}). Possible out-of-bounds write."
            )
    except ValueError:
        raise
    except Exception:
        pass  # openpyxl will validate the ref itself


def get_ledger_summary() -> str:
    """Compact text summary of the ledger for the AI prompt (30-second cache)."""
    now = time.time()
    if _summary_cache["data"] and now - _summary_cache["ts"] < 30:
        return _summary_cache["data"]

    wb = get_workbook(data_only=True)
    lines = ["Workbook: ledger.xlsx", ""]

    for name in wb.sheetnames:
        ws = wb[name]
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        lines.append(f"--- Sheet: '{name}' ({max_row} rows x {max_col} cols) ---")
        headers = [str(ws.cell(1, c).value or "") for c in range(1, max_col + 1)]
        lines.append(f"  Columns: {' | '.join(headers)}")
        start = max(2, max_row - 7)
        for r in range(start, max_row + 1):
            vals = [str(ws.cell(r, c).value or "") for c in range(1, max_col + 1)]
            lines.append(f"  Row {r}: {' | '.join(vals)}")
        lines.append("")

    wb.close()
    summary = "\n".join(lines)
    _summary_cache["data"] = summary
    _summary_cache["ts"] = now
    return summary


# ── Monetary formatting & running balance ─────────────────────────

_MONEY_FORMAT = "#,##0.00"
_MONETARY_HEADERS = ("debit", "credit", "balance", "amount")


def _monetary_columns(ws) -> set[int]:
    """1-based indices of columns whose header looks monetary."""
    cols: set[int] = set()
    for c in range(1, (ws.max_column or 1) + 1):
        h = ws.cell(1, c).value
        if h and any(k in str(h).strip().lower() for k in _MONETARY_HEADERS):
            cols.add(c)
    return cols


def _apply_money_format(ws, row: int, col: int, value):
    """Apply currency number format if the cell holds a numeric value (item 3.2)."""
    if isinstance(value, (int, float)):
        ws.cell(row=row, column=col).number_format = _MONEY_FORMAT


def recalculate_running_balance(ws):
    """Recompute the GeneralLedger Balance column as a cumulative debit-credit
    running total down the sheet (item 1.2). No-op if columns are missing."""
    headers = {str(ws.cell(1, c).value).strip().lower(): c
               for c in range(1, (ws.max_column or 1) + 1)
               if ws.cell(1, c).value}
    c_debit = headers.get("debit")
    c_credit = headers.get("credit")
    c_balance = headers.get("balance")
    if not (c_debit and c_credit and c_balance):
        return
    running = 0.0
    for r in range(2, (ws.max_row or 1) + 1):
        debit = ws.cell(r, c_debit).value
        credit = ws.cell(r, c_credit).value
        try:
            running += float(debit or 0) - float(credit or 0)
        except (TypeError, ValueError):
            pass  # leave running unchanged on non-numeric rows
        ws.cell(row=r, column=c_balance, value=round(running, 2))
        ws.cell(r, c_balance).number_format = _MONEY_FORMAT


# ── Snapshot restore (undo) ───────────────────────────────────────

def restore_snapshot(snapshot_bytes: bytes):
    """Overwrite the ledger file with a previously stored snapshot (item 2.2)."""
    with _get_lock():
        _atomic_replace_with_bytes(snapshot_bytes)
    _invalidate_caches()
    logger.info("Ledger restored from snapshot (%d bytes).", len(snapshot_bytes))


# ── Double-entry pre-flight check ─────────────────────────────────

def _check_double_entry(actions: list[dict]):
    """Raise ValueError if the proposed actions violate double-entry balance."""
    total_debits = 0.0
    total_credits = 0.0

    # Locate the Debit/Credit columns by header name rather than fixed position,
    # so the check stays correct if the GeneralLedger columns are ever reordered.
    try:
        gl_cols = get_column_indices("GeneralLedger")
    except Exception:
        gl_cols = {}
    debit_idx = gl_cols.get("Debit", 6) - 1   # values[] is 0-based; columns are 1-based
    credit_idx = gl_cols.get("Credit", 7) - 1

    for action in actions:
        op = action.get("operation", "")
        ctx = action.get("context", "").lower()

        if op == "insert_row":
            vals = action.get("values") or []
            if len(vals) > max(debit_idx, credit_idx):
                try:
                    d = vals[debit_idx]
                    c = vals[credit_idx]
                    if isinstance(d, (int, float)) and d > 0:
                        total_debits += float(d)
                    if isinstance(c, (int, float)) and c > 0:
                        total_credits += float(c)
                except (IndexError, TypeError):
                    pass
        else:
            val = action.get("new_value")
            if isinstance(val, (int, float)):
                if "debit" in ctx:
                    total_debits += float(val)
                elif "credit" in ctx:
                    total_credits += float(val)

    if total_debits > 0 and total_credits > 0:
        if abs(total_debits - total_credits) > 0.01:
            raise ValueError(
                f"Double-entry violation: debits={total_debits:.2f} "
                f"!= credits={total_credits:.2f}"
            )


# ── Execute ───────────────────────────────────────────────────────

def execute_actions(actions: list[dict]) -> tuple[bytes, list[str]]:
    """Execute approved proposal actions under the file lock.

    Returns ``(pre_snapshot, change_log)``. ``pre_snapshot`` is the ledger's exact
    bytes *before* any write, captured inside the same lock that guards the write,
    so it is always a consistent basis for undo (see ``restore_snapshot``).
    """
    _check_double_entry(actions)

    change_log: list[str] = []

    modified_sheets: set[str] = set()

    with _get_lock():
        snapshot = take_snapshot()  # consistent pre-state, captured under the lock
        wb = get_workbook()
        valid_accounts = get_valid_accounts(wb)

        try:
            for action in actions:
                op = action["operation"]
                sheet = action["sheet"]

                if sheet not in wb.sheetnames:
                    raise ValueError(f"Sheet '{sheet}' does not exist.")

                ws = wb[sheet]
                modified_sheets.add(sheet)
                money_cols = _monetary_columns(ws)

                if op == "write_cell":
                    cell_ref = action["cell_ref"]
                    _validate_cell_bounds(ws, cell_ref)
                    old_val = ws[cell_ref].value
                    new_val = action["new_value"]
                    if isinstance(new_val, float):
                        new_val = round(new_val, 2)
                    ws[cell_ref] = new_val
                    col_letter, row_num = coordinate_from_string(cell_ref)
                    if column_index_from_string(col_letter) in money_cols:
                        _apply_money_format(ws, row_num, column_index_from_string(col_letter), new_val)
                    change_log.append(
                        f"[{sheet}!{cell_ref}] {old_val!r} → {new_val!r}"
                        + (f" ({action.get('context', '')})" if action.get("context") else "")
                    )

                elif op == "write_formula":
                    cell_ref = action["cell_ref"]
                    _validate_cell_bounds(ws, cell_ref)
                    formula = action["formula"]
                    ws[cell_ref] = formula
                    change_log.append(f"[{sheet}!{cell_ref}] formula: {formula}")

                elif op == "write_range":
                    start = action["start_cell"]
                    end = action["end_cell"]
                    values = action["values_2d"]
                    from openpyxl.utils import range_boundaries
                    min_col, min_row, _, _ = range_boundaries(f"{start}:{end}")
                    for r_idx, row_data in enumerate(values):
                        for c_idx, val in enumerate(row_data):
                            if isinstance(val, float):
                                val = round(val, 2)
                            tr, tc = min_row + r_idx, min_col + c_idx
                            ws.cell(row=tr, column=tc, value=val)
                            if tc in money_cols:
                                _apply_money_format(ws, tr, tc, val)
                    change_log.append(
                        f"[{sheet}!{start}:{end}] wrote "
                        f"{len(values)}×{len(values[0]) if values else 0} block"
                    )

                elif op == "insert_row":
                    row_idx = action["row_index"]
                    values = action["values"]
                    # Validate account number (column D = 0-based index 3) against ChartOfAccounts
                    if valid_accounts and len(values) > 3 and values[3] is not None:
                        acct = str(values[3]).strip()
                        if acct and acct not in valid_accounts:
                            raise ValueError(
                                f"Account '{acct}' does not exist in the Chart of Accounts. "
                                f"Valid accounts: {', '.join(sorted(valid_accounts))}"
                            )
                    ws.insert_rows(row_idx)
                    for c_idx, val in enumerate(values, start=1):
                        if isinstance(val, float):
                            val = round(val, 2)
                        ws.cell(row=row_idx, column=c_idx, value=val)
                        if c_idx in money_cols:
                            _apply_money_format(ws, row_idx, c_idx, val)
                    change_log.append(
                        f"[{sheet}] inserted row {row_idx} ({len(values)} values)"
                    )

                else:
                    raise ValueError(f"Unknown operation: {op}")

            # Recompute the running balance on any GeneralLedger we touched (item 1.2)
            if "GeneralLedger" in modified_sheets:
                recalculate_running_balance(wb["GeneralLedger"])

            tmp_path = LEDGER_PATH + ".tmp"
            wb.save(tmp_path)
            os.replace(tmp_path, LEDGER_PATH)  # atomic swap: never a torn .xlsx
            logger.info("Ledger saved — %d action(s) executed.", len(actions))
            _invalidate_caches()

        finally:
            wb.close()

    return snapshot, change_log
