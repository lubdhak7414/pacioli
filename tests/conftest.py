"""Shared pytest fixtures: a known in-memory ledger and isolated DB path."""

import openpyxl
import pytest

import ledger_engine


def _build_ledger(path):
    """Create a small, known ledger:

    TXN1  Cash +3500 / Service Revenue +3500
    TXN2  Office Supplies +1200 / Cash -1200
    TXN3  Rent Expense +2000 / Cash -2000

    => Cash net debit 300, Revenue 3500, Expenses 3200, Net income 300.
    """
    wb = openpyxl.Workbook()
    gl = wb.active
    gl.title = "GeneralLedger"
    gl.append(["Date", "Ref", "Description", "Account", "AccountName",
               "Debit", "Credit", "Balance"])
    gl.append(["", "", "Opening Balance", "", "", 0, 0, 0])
    rows = [
        ("2026-01-01", "TXN-001", "Client payment", "1010", "Cash", 3500, 0),
        ("2026-01-01", "TXN-001", "Client payment", "4100", "Service Revenue", 0, 3500),
        ("2026-01-02", "TXN-002", "Supplies", "6050", "Office Supplies", 1200, 0),
        ("2026-01-02", "TXN-002", "Supplies", "1010", "Cash", 0, 1200),
        ("2026-01-03", "TXN-003", "Rent", "6100", "Rent Expense", 2000, 0),
        ("2026-01-03", "TXN-003", "Rent", "1010", "Cash", 0, 2000),
    ]
    for d, ref, desc, acct, name, dr, cr in rows:
        gl.append([d, ref, desc, acct, name, dr, cr, 0])

    tb = wb.create_sheet("TrialBalance")
    tb.append(["Account", "AccountName", "Debit", "Credit"])

    coa = wb.create_sheet("ChartOfAccounts")
    coa.append(["Account", "AccountName", "Type"])
    coa.append(["1010", "Cash", "Asset"])
    coa.append(["1200", "Accounts Receivable", "Asset"])
    coa.append(["2100", "Accounts Payable", "Liability"])
    coa.append(["3100", "Retained Earnings", "Equity"])
    coa.append(["4100", "Service Revenue", "Revenue"])
    coa.append(["6050", "Office Supplies", "Expense"])
    coa.append(["6100", "Rent Expense", "Expense"])
    wb.save(path)


@pytest.fixture
def ledger(tmp_path, monkeypatch):
    """Point ledger_engine at a fresh known workbook and clear its caches."""
    path = tmp_path / "ledger.xlsx"
    _build_ledger(str(path))
    monkeypatch.setattr(ledger_engine, "LEDGER_PATH", str(path))
    ledger_engine._invalidate_caches()
    yield str(path)
    ledger_engine._invalidate_caches()
