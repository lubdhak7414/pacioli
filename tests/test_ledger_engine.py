"""Execution safety nets: account validation, double-entry, running balance, formatting."""

import os

import openpyxl
import pytest

import ledger_engine


def _row(row_index, account, name, debit, credit):
    return {
        "operation": "insert_row", "sheet": "GeneralLedger", "row_index": row_index,
        "values": ["2026-01-05", "TXN-NEW", "desc", account, name, debit, credit, 0],
    }


def test_running_balance_recomputed(ledger):
    ledger_engine.execute_actions([
        _row(9, "1010", "Cash", 500, 0),
        _row(10, "1200", "Accounts Receivable", 0, 500),
    ])
    wb = openpyxl.load_workbook(ledger)
    gl = wb["GeneralLedger"]
    assert gl["H3"].value == 3500.0   # first txn cumulative
    assert gl["H4"].value == 0.0      # offsetting credit
    assert gl["H9"].value == 500.0    # new debit
    assert gl["H10"].value == 0.0     # new offsetting credit
    wb.close()


def test_currency_format_applied(ledger):
    ledger_engine.execute_actions([
        _row(9, "1010", "Cash", 500, 0),
        _row(10, "1200", "Accounts Receivable", 0, 500),
    ])
    wb = openpyxl.load_workbook(ledger)
    gl = wb["GeneralLedger"]
    assert gl["F9"].number_format == "#,##0.00"
    assert gl["H3"].number_format == "#,##0.00"
    wb.close()


def test_unknown_account_rejected(ledger):
    with pytest.raises(ValueError, match="Chart of Accounts"):
        ledger_engine.execute_actions([_row(9, "9999", "Bogus", 100, 0)])


def test_double_entry_violation_rejected(ledger):
    with pytest.raises(ValueError, match="[Dd]ouble-entry"):
        ledger_engine.execute_actions([
            _row(9, "1010", "Cash", 500, 0),
            _row(10, "1200", "Accounts Receivable", 0, 400),
        ])


def test_valid_accounts_loaded(ledger):
    accts = ledger_engine.get_valid_accounts()
    assert "1010" in accts and "4100" in accts
    assert "9999" not in accts


def test_restore_snapshot_roundtrip(ledger):
    before = ledger_engine.take_snapshot()
    ledger_engine.execute_actions([
        _row(9, "1010", "Cash", 500, 0),
        _row(10, "1200", "Accounts Receivable", 0, 500),
    ])
    wb = openpyxl.load_workbook(ledger)
    assert wb["GeneralLedger"].max_row == 10
    wb.close()

    ledger_engine.restore_snapshot(before)
    wb = openpyxl.load_workbook(ledger)
    assert wb["GeneralLedger"].max_row == 8   # back to original
    wb.close()


def test_execute_returns_restorable_snapshot(ledger):
    """execute_actions returns the pre-execution snapshot, captured under the lock."""
    snapshot, change_log = ledger_engine.execute_actions([
        _row(9, "1010", "Cash", 500, 0),
        _row(10, "1200", "Accounts Receivable", 0, 500),
    ])
    assert isinstance(snapshot, bytes) and change_log

    wb = openpyxl.load_workbook(ledger)
    assert wb["GeneralLedger"].max_row == 10
    wb.close()

    # Restoring the returned snapshot undoes the change — proving it is pre-state.
    ledger_engine.restore_snapshot(snapshot)
    wb = openpyxl.load_workbook(ledger)
    assert wb["GeneralLedger"].max_row == 8
    wb.close()


def test_execute_leaves_no_temp_file(ledger):
    """The atomic write swaps in via .tmp; it must not linger afterwards."""
    ledger_engine.execute_actions([
        _row(9, "1010", "Cash", 500, 0),
        _row(10, "1200", "Accounts Receivable", 0, 500),
    ])
    assert not os.path.exists(ledger + ".tmp")


def test_get_chart_text_lists_accounts(ledger):
    text = ledger_engine.get_chart_text()
    assert "1010 Cash (Asset)" in text
    assert "4100 Service Revenue (Revenue)" in text
